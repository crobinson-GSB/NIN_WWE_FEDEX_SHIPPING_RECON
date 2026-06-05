import streamlit as st
import pandas as pd
import re
import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="GSB Digital — Shipping Reconciliation",
    page_icon="📦",
    layout="centered"
)

# ── Error code system ──
# Every failure point raises a ReconError with a code, plain-English title,
# and actionable guidance. The run block catches these and shows a structured
# error card instead of a raw Python traceback.

class ReconError(Exception):
    def __init__(self, code, title, detail, fix):
        self.code   = code    # e.g. "VIS-001"
        self.title  = title   # short plain-English description
        self.detail = detail  # what specifically went wrong
        self.fix    = fix     # what the user should do
        super().__init__(title)

def show_error(e):
    if isinstance(e, ReconError):
        st.markdown(f"""
<div style='background:#FFF0F0; border-left:4px solid #CC0000; border-radius:6px; padding:1rem 1.25rem; margin:1rem 0;'>
<div style='font-size:11px; color:#999; font-family:monospace; margin-bottom:4px;'>ERROR {e.code}</div>
<div style='font-size:15px; font-weight:600; color:#CC0000; margin-bottom:6px;'>{e.title}</div>
<div style='font-size:13px; color:#2B2B2B; margin-bottom:10px;'>{e.detail}</div>
<div style='font-size:13px; background:#fff; border-radius:4px; padding:8px 12px; border:1px solid #FFCCCC;'>
<strong>What to do:</strong> {e.fix}
</div>
</div>
""", unsafe_allow_html=True)
    else:
        # Unexpected error — show code and a generic message
        import traceback
        tb_str = traceback.format_exc()
        st.markdown(f"""
<div style='background:#FFF0F0; border-left:4px solid #CC0000; border-radius:6px; padding:1rem 1.25rem; margin:1rem 0;'>
<div style='font-size:11px; color:#999; font-family:monospace; margin-bottom:4px;'>ERROR SYS-999 — UNEXPECTED</div>
<div style='font-size:15px; font-weight:600; color:#CC0000; margin-bottom:6px;'>Something unexpected went wrong</div>
<div style='font-size:13px; color:#2B2B2B; margin-bottom:10px;'>The tool hit an error it didn't anticipate. The technical details are below.</div>
<div style='font-size:13px; background:#fff; border-radius:4px; padding:8px 12px; border:1px solid #FFCCCC;'>
<strong>What to do:</strong> Take a screenshot of this error and send it to whoever maintains this tool. Include the files you were trying to upload.
</div>
<details><summary style='font-size:12px; color:#999; margin-top:8px; cursor:pointer;'>Technical details (click to expand)</summary>
<pre style='font-size:11px; color:#666; margin-top:6px; white-space:pre-wrap;'>{str(e)}\n\n{tb_str}</pre>
</details>
</div>
""", unsafe_allow_html=True)

st.markdown("""
<style>
    .stApp { background-color: #f9f9f9; }
    .block-container { max-width: 760px; padding-top: 2rem; }
    h1 { color: #E8601C !important; font-size: 1.6rem !important; }
    h3 { color: #2B2B2B !important; font-size: 1rem !important; font-weight: 600 !important; }
    p { color: #2B2B2B !important; }
    div[data-testid="stFileUploader"] {
        background-color: #ffffff !important;
        border: 1.5px dashed #CCCCCC !important;
        border-radius: 8px !important;
        padding: 0.5rem !important;
    }
    div[data-testid="stFileUploader"]:hover {
        border-color: #E8601C !important;
    }
    div[data-testid="stFileUploader"] label {
        font-weight: 600;
        color: #2B2B2B !important;
    }
    div[data-testid="stFileUploader"] small {
        color: #777777 !important;
    }
    div[data-testid="stMarkdownContainer"] h3 {
        border-left: 4px solid #E8601C;
        padding-left: 10px;
    }
    div[data-testid="stExpander"] {
        border: 2px solid #E8601C !important;
        border-radius: 6px !important;
    }
    div[data-testid="stButton"] {
        display: flex;
        justify-content: center;
    }
    .stButton button {
        background-color: #E8601C !important;
        color: white !important;
        border: none !important;
        font-weight: 600 !important;
        padding: 0.75rem 2rem !important;
        border-radius: 6px !important;
        width: 100% !important;
        font-size: 1rem !important;
    }
    .stButton button:hover { background-color: #c94f14 !important; }
    div[data-testid="stAlert"] {
        background-color: #FDF3EE !important;
        border: 1px solid #E8601C !important;
        border-radius: 6px !important;
        color: #2B2B2B !important;
    }
    div[data-testid="stFileUploader"] section {
        background-color: #2B2B2B !important;
        border-color: #444444 !important;
    }
    div[data-testid="stFileUploader"] section:hover {
        border-color: #E8601C !important;
    }
    div[data-testid="stFileUploader"] section button {
        background-color: #E8601C !important;
        color: white !important;
        border: none !important;
    }
    div[data-testid="stFileUploader"] section button:hover {
        background-color: #c94f14 !important;
    }
    div[data-testid="stFileUploader"] section span {
        color: #AAAAAA !important;
    }
    footer { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

st.markdown("<div style='padding-top: 2rem;'>", unsafe_allow_html=True)
st.image("https://gsbdigital.com/wp-content/uploads/2018/06/GSB_DigitalLogo-2015-300x44.png", width=200)
st.markdown("</div>", unsafe_allow_html=True)
st.markdown("# Shipping Reconciliation Tool")
st.markdown("Automatically match Vision exports against vendor invoices and identify shipping cost discrepancies.")

with st.expander("How to use this tool", expanded=False):
    st.markdown("""
**What this tool does**
Compares your Printsmith Vision shipping export against vendor invoices from NIN (courier) and WWE/UPS. It cleans and matches invoice numbers automatically, then flags any discrepancies between what was recorded in Vision and what the vendor actually charged.

**Step 1 — Export from Vision**
In Printsmith Vision, run your Digital Shipping report for the billing period. Export it as a .txt or .xlsx file.

**Step 2 — Download vendor invoices**
- **NIN:** Download your courier invoice from the NIN portal as an .xls file
- **WWE/UPS:** Download your UPS invoice from the WWE portal as an .xls file
- **FedEx:** Download your FedEx invoice if applicable

**Step 3 — Set the billing period**
Enter the start and end dates for the period you are reconciling. This will be used to name the downloaded report.

**Step 4 — Upload the files**
Drop each file into the correct upload box below. The Vision export and at least one vendor invoice (NIN or WWE) are required.

**Step 5 — Run the reconciliation**
Click "Run Reconciliation." The tool will process the files and show a summary of results.

**Step 6 — Download the report**
Click the download button to get your Excel report. It includes five tabs:
- **Summary** — overall counts, full financial totals, and mismatch impact broken out separately
- **Mismatches** — invoices where Vision cost and vendor cost don't match (action required)
- **Matched** — invoices that reconciled cleanly
- **Not in Vision** — vendor charges with no matching Vision entry
- **Vision Only** — Vision entries with no vendor invoice received yet
- **Possible Duplicates** — rows appearing in more than one uploaded file for the same vendor, flagged for your team to review

**Tips**
- Invoice numbers are matched automatically — no manual cleaning needed
- You can upload multiple files per vendor — useful when a billing period spans more than one invoice
- If files overlap in date range, duplicate shipments are flagged in the Possible Duplicates tab — nothing is deleted
- Run this each billing cycle when vendor invoices arrive
- The Mismatches tab is your primary action list
""")

st.markdown("<div style='margin-top: 0.25rem;'></div>", unsafe_allow_html=True)

# --- Billing period date range ---
st.markdown("### Billing Period")
dcol1, dcol2 = st.columns(2)
with dcol1:
    period_start = st.date_input("Period Start", value=None, key="period_start")
with dcol2:
    period_end = st.date_input("Period End", value=None, key="period_end")

st.markdown("### Vision Export — Required")
vision_file = st.file_uploader(
    "Vision Report (.txt or .xlsx export from Printsmith Vision)",
    type=["txt", "xlsx"],
    key="vision"
)

st.markdown("### Vendor Invoices")
st.caption("Drop all vendor files here — NIN, WWE, FedEx, and Gelato are detected automatically. You can upload multiple files per vendor.")
vendor_files = st.file_uploader(
    "Vendor invoice files (NIN, WWE/UPS, FedEx, Gelato)",
    type=["xls", "xlsx", "csv"],
    key="vendors",
    accept_multiple_files=True
)

st.markdown("<hr style='border: none; border-top: 1px solid #DDDDDD; margin: 1.5rem 0;'>", unsafe_allow_html=True)

# --- Helpers ---

def clean_key(val):
    if pd.isna(val):
        return ''
    s = str(val).strip()
    s = re.sub(r'\.0$', '', s)
    s = re.sub(r'^[Dd]', '', s)
    s = s.split('/')[0]
    s = s.replace('-', '').replace(' ', '').strip()
    return s.upper()

def read_excel_any(file):
    """Read .xls or .xlsx without assuming engine."""
    try:
        if file.name.lower().endswith('.xlsx'):
            return pd.read_excel(file)
        else:
            return pd.read_excel(file, engine='xlrd')
    except Exception as ex:
        raise ReconError(
            code="FILE-001",
            title=f"Could not open file: {file.name}",
            detail=f"The file could not be read. Error: {ex}",
            fix="Make sure the file isn't open in Excel or another program, isn't password protected, "
                "and is a valid .xls or .xlsx file. Try re-downloading it from the vendor portal."
        )

REQUIRED_VISION  = {'Invoice', 'Cost', 'Amount', 'Sales Rep', 'Description', 'Pickup Date'}
# ── Fuzzy column matching ──
# Each entry is a list of keywords. A column matches if it contains ALL keywords
# (case-insensitive). The tool finds the best matching actual column for each
# required field, so minor vendor renames (e.g. adding "(Default)", changing
# spaces to underscores) are handled automatically.

FUZZY_VISION  = {
    'Invoice':      ['invoice'],
    'Cost':         ['cost'],
    'Amount':       ['amount'],
    'Sales Rep':    ['sales', 'rep'],
    'Description':  ['description'],
    'Pickup Date':  ['pickup', 'date'],
}
FUZZY_NIN = {
    'Auth':          ['auth'],
    'AmountCharged': ['amount', 'charge'],
    'InvoiceNumber': ['invoice', 'number'],
    'OrderNumber':   ['order', 'number'],
    'Orderdate':     ['order', 'date'],
}
FUZZY_WWE = {
    'Billing Reference 1': ['billing', 'reference'],
    'Charge Total':        ['charge', 'total'],
    'Invoice #':           ['invoice'],
    'Airbill #':           ['airbill'],
    'Ship date':           ['ship', 'date'],
}
FUZZY_FEDEX = {
    'Reference Notes Line 1':     ['reference', 'notes'],
    'Net Charge Amount USD':      ['net', 'charge'],
    'Shipment Tracking Number':   ['tracking', 'number'],
    'Invoice Number':             ['invoice', 'number'],
    'Shipment Date (mm/dd/yyyy)': ['shipment', 'date'],
}
FUZZY_GELATO = {
    'Packages Order Number':              ['order', 'number'],
    'Packages Tracking Number':           ['tracking', 'number'],
    'Packages First Scan Date':           ['scan', 'date'],
    'Packages Gross Transaction Value':   ['gross', 'transaction', 'value'],
    'Packages Carrier':                   ['carrier'],
}
FUZZY_GELATO2 = {
    'orderReferenceId': ['reference'],
    'costTotal':        ['cost'],
    'trackingNumber':   ['tracking'],
    'orderDate':        ['date'],
}

def fuzzy_find_col(df_cols, keywords):
    """Find the first column whose name contains all keywords (case-insensitive)."""
    cols_lower = [c.lower() for c in df_cols]
    for i, col_lower in enumerate(cols_lower):
        if all(kw.lower() in col_lower for kw in keywords):
            return df_cols[i]
    return None

def build_col_map(df, fuzzy_spec):
    """
    Build a mapping from canonical field name → actual column name in df.
    Returns (col_map, missing) where missing is a list of fields not found.
    """
    col_map = {}
    missing = []
    df_cols = list(df.columns)
    for canonical, keywords in fuzzy_spec.items():
        found = fuzzy_find_col(df_cols, keywords)
        if found:
            col_map[canonical] = found
        else:
            missing.append(canonical)
    return col_map, missing

def fuzzy_detect_vendor(df):
    """Identify which vendor a dataframe belongs to using fuzzy column matching."""
    scores = {}
    for vendor, spec in [('vision', FUZZY_VISION), ('nin', FUZZY_NIN),
                          ('wwe', FUZZY_WWE), ('fedex', FUZZY_FEDEX),
                          ('gelato', FUZZY_GELATO), ('gelato2', FUZZY_GELATO2)]:
        _, missing = build_col_map(df, spec)
        scores[vendor] = len(spec) - len(missing)
    best = max(scores, key=scores.get)
    best_score = scores[best]
    total = len({'vision': FUZZY_VISION, 'nin': FUZZY_NIN, 'wwe': FUZZY_WWE,
                 'fedex': FUZZY_FEDEX, 'gelato': FUZZY_GELATO,
                 'gelato2': FUZZY_GELATO2}[best if best != 'gelato2' else 'gelato2'])
    # Must match at least 60% of expected columns to be confident
    if best_score / total < 0.6:
        return 'unknown', {}
    col_map, _ = build_col_map(df, {'vision': FUZZY_VISION, 'nin': FUZZY_NIN,
                                     'wwe': FUZZY_WWE, 'fedex': FUZZY_FEDEX,
                                     'gelato': FUZZY_GELATO,
                                     'gelato2': FUZZY_GELATO2}[best if best != 'gelato2' else 'gelato2'])
    return ('gelato' if best == 'gelato2' else best), col_map

def validate_columns(df, fuzzy_spec, label):
    """Fuzzy validation — raises ReconError if too many columns missing."""
    col_map, missing = build_col_map(df, fuzzy_spec)
    if len(missing) > len(fuzzy_spec) * 0.4:
        raise ReconError(
            code="COL-001",
            title=f"Unrecognized file format — {label}",
            detail=f"The file doesn't look like a {label} export. "
                   f"Could not find columns matching: {', '.join(sorted(missing))}. "
                   f"Columns found in the file: {', '.join(list(df.columns)[:8])}{'...' if len(df.columns) > 8 else ''}.",
            fix=f"Make sure you're uploading the correct {label} export file. "
                f"If the vendor recently changed their export format, the column names "
                f"may have changed enough that the tool can't recognize them. "
                f"Contact whoever maintains this tool and share the file."
        )
    return col_map

# Keep these for detect_vendor file-level detection (reads only header row)
REQUIRED_NIN    = {'Auth', 'AmountCharged', 'InvoiceNumber', 'OrderNumber'}
REQUIRED_WWE    = {'Billing Reference 1', 'Charge Total', 'Invoice #', 'Airbill #'}
REQUIRED_FEDEX  = {'Reference Notes Line 1', 'Net Charge Amount USD', 'Shipment Tracking Number', 'Invoice Number'}
REQUIRED_GELATO = {'Packages Order Number', 'Packages Tracking Number', 'Packages First Scan Date'}
REQUIRED_GELATO2= {'orderReferenceId', 'costTotal', 'trackingNumber', 'orderDate'}

def load_vision(file):
    try:
        if file.name.endswith('.txt'):
            try:
                df = pd.read_csv(file, sep='\t', encoding='cp1252')
            except UnicodeDecodeError:
                file.seek(0)
                df = pd.read_csv(file, sep='\t')
        else:
            df = pd.read_excel(file)
    except Exception as ex:
        raise ReconError(
            code="VIS-001",
            title="Could not open Vision file",
            detail=f"The Vision file '{file.name}' could not be read. Error: {ex}",
            fix="Make sure the file isn't open in another program (like Excel), isn't corrupted, "
                "and is a .txt or .xlsx export from Printsmith Vision."
        )
    if df.empty:
        raise ReconError(
            code="VIS-002",
            title="Vision file is empty",
            detail=f"The file '{file.name}' opened successfully but contains no data rows.",
            fix="Re-export the Vision report and make sure the date range contains shipments."
        )
    col_map = validate_columns(df, FUZZY_VISION, 'Vision Export')
    cost_col   = col_map.get('Cost', 'Cost')
    amount_col = col_map.get('Amount', 'Amount')
    inv_col    = col_map.get('Invoice', 'Invoice')
    rep_col    = col_map.get('Sales Rep', 'Sales Rep')
    desc_col   = col_map.get('Description', 'Description')
    date_col   = col_map.get('Pickup Date', 'Pickup Date')
    df[cost_col]   = pd.to_numeric(df[cost_col],   errors='coerce').fillna(0)
    df[amount_col] = pd.to_numeric(df[amount_col], errors='coerce').fillna(0)
    if df[cost_col].sum() == 0:
        raise ReconError(
            code="VIS-003",
            title="Vision file has no cost data",
            detail=f"The Cost column in '{file.name}' is present but all values are zero or blank.",
            fix="Check that you exported the correct report from Vision. The Digital Shipping report "
                "should have cost values populated. If costs are genuinely zero, the reconciliation "
                "will still run but all vendor charges will appear as mismatches."
        )
    df['key']           = df[inv_col].apply(clean_key)
    df['Invoice_clean'] = df['key']
    agg = df.groupby('key').agg(
        vision_invoice=    ('Invoice_clean', 'first'),
        vision_invoice_raw=(inv_col,         'first'),
        vision_sales_rep=  (rep_col,         'first'),
        vision_billed=     (amount_col,      'sum'),
        vision_cost=       (cost_col,        'sum'),
        vision_description=(desc_col,        'first'),
        vision_date=       (date_col,        'first'),
    ).reset_index()
    return agg, df

def assign_unmatched_keys(df, key_col, label):
    """
    For rows where key is blank or non-standard (not a 6-7 digit GSB invoice number),
    assign a unique placeholder key so they are NEVER dropped from the reconciliation.
    Each unmatched row gets its own key: e.g. NIN_UNMATCHED_001, WWE_UNMATCHED_002
    This guarantees every dollar from every vendor file appears in the output.
    """
    gsb_pattern = re.compile(r'^\d{6,7}$')
    counter = [0]
    def safe_key(k):
        if k == '' or not gsb_pattern.match(k):
            counter[0] += 1
            return f'{label}_UNMATCHED_{counter[0]:03d}'
        return k
    df[key_col] = df[key_col].apply(safe_key)
    return df

def load_nin(file):
    df = read_excel_any(file)
    if df.empty:
        raise ReconError("NIN-001", "NIN file is empty",
                         f"'{file.name}' contains no data rows.",
                         "Re-download the NIN invoice from the portal and try again.")
    col_map = validate_columns(df, FUZZY_NIN, 'NIN — Courier')
    auth_col   = col_map.get('Auth', 'Auth')
    amount_col = col_map.get('AmountCharged', 'AmountCharged')
    inv_col    = col_map.get('InvoiceNumber', 'InvoiceNumber')
    order_col  = col_map.get('OrderNumber', 'OrderNumber')
    date_col   = col_map.get('Orderdate', 'Orderdate')
    df[amount_col] = pd.to_numeric(df[amount_col], errors='coerce').fillna(0)
    df['key'] = df[auth_col].apply(clean_key)
    df = assign_unmatched_keys(df, 'key', 'NIN')
    df['_tracking']    = df[order_col].astype(str).str.strip()
    df['_source_file'] = file.name
    df['_amount']  = df[amount_col]
    df['_inv_num'] = df[inv_col]  if inv_col  in df.columns else ''
    df['_order']   = df[order_col]
    df['_auth']    = df[auth_col]
    df['_date']    = df[date_col] if date_col in df.columns else None
    return df

def load_wwe(file):
    df = read_excel_any(file)
    if df.empty:
        raise ReconError("WWE-001", "WWE file is empty",
                         f"'{file.name}' contains no data rows.",
                         "Re-download the WWE/UPS invoice from the portal and try again.")
    col_map = validate_columns(df, FUZZY_WWE, 'WWE — UPS')
    ref_col     = col_map.get('Billing Reference 1', 'Billing Reference 1')
    charge_col  = col_map.get('Charge Total', 'Charge Total')
    inv_col     = col_map.get('Invoice #', 'Invoice #')
    airbill_col = col_map.get('Airbill #', 'Airbill #')
    date_col    = col_map.get('Ship date', 'Ship date')
    df[charge_col] = pd.to_numeric(df[charge_col], errors='coerce').fillna(0)
    df['key'] = df[ref_col].apply(clean_key)
    df['key'] = df['key'].apply(lambda s: re.match(r'^(\d{6,7})', s).group(1)
                                if re.match(r'^(\d{6,7})', s) else s)
    df = assign_unmatched_keys(df, 'key', 'WWE')
    df['_tracking']    = df[airbill_col].astype(str).str.strip()
    df['_source_file'] = file.name
    df['_amount']  = df[charge_col]
    df['_inv_num'] = df[inv_col]     if inv_col     in df.columns else ''
    df['_airbill'] = df[airbill_col]
    df['_ref']     = df[ref_col]
    df['_date']    = df[date_col]    if date_col    in df.columns else None
    return df

def load_fedex(file):
    if file.name.lower().endswith('.csv'):
        try:
            df = pd.read_csv(file, encoding='cp1252')
        except UnicodeDecodeError:
            file.seek(0)
            df = pd.read_csv(file)
    else:
        df = read_excel_any(file)
    col_map = validate_columns(df, FUZZY_FEDEX, 'FedEx')
    if df.empty:
        raise ReconError("FDX-001", "FedEx file is empty",
                         f"'{file.name}' contains no data rows.",
                         "Re-download the FedEx invoice and try again.")
    ref_col      = col_map.get('Reference Notes Line 1', 'Reference Notes Line 1')
    charge_col   = col_map.get('Net Charge Amount USD', 'Net Charge Amount USD')
    tracking_col = col_map.get('Shipment Tracking Number', 'Shipment Tracking Number')
    inv_col      = col_map.get('Invoice Number', 'Invoice Number')
    date_col     = col_map.get('Shipment Date (mm/dd/yyyy)', 'Shipment Date (mm/dd/yyyy)')
    df[charge_col] = pd.to_numeric(df[charge_col], errors='coerce').fillna(0)
    df['key'] = df[ref_col].apply(lambda v: clean_key(v).split('_')[0])
    df = assign_unmatched_keys(df, 'key', 'FEDEX')
    df['_tracking']    = df[tracking_col].astype(str).str.strip()
    df['_source_file'] = file.name
    df['_amount']  = df[charge_col]
    df['_inv_num'] = df[inv_col]      if inv_col      in df.columns else ''
    df['_ref']     = df[ref_col]
    df['_date']    = df[date_col]     if date_col     in df.columns else None
    return df

def load_gelato(file):
    if file.name.lower().endswith('.csv'):
        try:
            df = pd.read_csv(file, encoding='cp1252')
        except UnicodeDecodeError:
            file.seek(0)
            df = pd.read_csv(file)
    else:
        df = read_excel_any(file)
    # Score both Gelato formats and use whichever matches better
    _, missing_excel = build_col_map(df, FUZZY_GELATO)
    _, missing_csv   = build_col_map(df, FUZZY_GELATO2)
    if len(missing_excel) <= len(missing_csv):
        col_map = validate_columns(df, FUZZY_GELATO, 'Gelato')
        order_col    = col_map.get('Packages Order Number', 'Packages Order Number')
        tracking_col = col_map.get('Packages Tracking Number', 'Packages Tracking Number')
        date_col     = col_map.get('Packages First Scan Date', 'Packages First Scan Date')
        cost_col     = col_map.get('Packages Gross Transaction Value')
        carrier_col  = col_map.get('Packages Carrier')
        if cost_col is None:
            raise ReconError(
                code="GEL-002",
                title="Gelato file missing cost column",
                detail=f"Could not find a column containing 'gross', 'transaction', and 'value' in '{file.name}'. "
                       f"Columns found: {', '.join(list(df.columns)[:8])}...",
                fix="The Gelato export format may have changed. Re-download the file and try again. "
                    "If the problem persists, contact whoever maintains this tool."
            )
        df[cost_col] = pd.to_numeric(df[cost_col], errors='coerce').fillna(0)
        df['key'] = df[order_col].apply(clean_key)
        df = assign_unmatched_keys(df, 'key', 'GELATO')
        df['_tracking']    = df[tracking_col].astype(str).str.strip()
        df['_source_file'] = file.name
        df['_cost']    = df[cost_col]
        df['_date']    = df[date_col]
        df['_ref']     = df[order_col].astype(str)
        df['_carrier'] = df[carrier_col] if carrier_col else 'Gelato'
    else:
        col_map = validate_columns(df, FUZZY_GELATO2, 'Gelato (CSV)')
        ref_col      = col_map.get('orderReferenceId', 'orderReferenceId')
        cost_col     = col_map.get('costTotal', 'costTotal')
        tracking_col = col_map.get('trackingNumber', 'trackingNumber')
        date_col     = col_map.get('orderDate', 'orderDate')
        df[cost_col] = pd.to_numeric(df[cost_col], errors='coerce').fillna(0)
        def extract_gsb(val):
            if pd.isna(val): return ''
            import re as _re
            m = _re.match(r'^(\d{6,7})', str(val).strip())
            return m.group(1) if m else clean_key(val)
        df['key'] = df[ref_col].apply(extract_gsb)
        df = assign_unmatched_keys(df, 'key', 'GELATO')
        df['_tracking']    = df[tracking_col].astype(str).str.strip()
        df['_source_file'] = file.name
        df['_cost']    = df[cost_col]
        df['_date']    = df[date_col]
        df['_ref']     = df[ref_col].astype(str)
        carrier_col    = 'shippingMethodCarrier'
        df['_carrier'] = df[carrier_col] if carrier_col in df.columns else 'Gelato'
    return df
def agg_gelato(df):
    return df.groupby('key').agg(
        gelato_actual_cost=('_cost', 'sum'),
        gelato_shipments=('_tracking', 'count'),
        gelato_raw_ref=('_ref', 'first'),
        gelato_carrier=('_carrier', 'first'),
        gelato_date=('_date', 'first')
    ).reset_index()

def detect_vendor(file):
    """Read column headers and identify vendor using fuzzy matching."""
    try:
        if file.name.lower().endswith('.csv'):
            try:
                df = pd.read_csv(file, nrows=1, encoding='cp1252')
            except UnicodeDecodeError:
                file.seek(0)
                df = pd.read_csv(file, nrows=1)
        else:
            try:
                df = pd.read_excel(file, nrows=1)
            except Exception:
                file.seek(0)
                df = pd.read_excel(file, nrows=1, engine='xlrd')
        file.seek(0)
        best_vendor = 'unknown'
        best_score  = 0
        for vendor, spec in [('nin', FUZZY_NIN), ('wwe', FUZZY_WWE),
                              ('fedex', FUZZY_FEDEX), ('gelato', FUZZY_GELATO),
                              ('gelato', FUZZY_GELATO2)]:
            _, missing = build_col_map(df, spec)
            score = (len(spec) - len(missing)) / len(spec)
            if score > best_score:
                best_score  = score
                best_vendor = vendor
        return best_vendor if best_score >= 0.6 else 'unknown'
    except Exception:
        file.seek(0)
        return 'unknown'

def combine_vendor_files(file_list, load_fn):
    """Load multiple files for the same vendor and combine into one dataframe.
    Keeps all rows — duplicates are flagged, never deleted."""
    frames = []
    for f in file_list:
        df = load_fn(f)
        if df is not None:
            frames.append(df)
    if not frames:
        return None
    return pd.concat(frames, ignore_index=True)

def flag_duplicates(df, tracking_col, cost_col, date_col, ref_col, label):
    """
    Flag rows where the same tracking number appears more than once.
    Returns the full dataframe (nothing removed) plus a separate
    duplicates dataframe for the Possible Duplicates tab.
    """
    df = df.copy()
    df['_possible_duplicate'] = ''
    df['_dup_files'] = ''

    valid = df[tracking_col].notna() & (df[tracking_col].astype(str).str.strip() != '') & \
            (df[tracking_col].astype(str).str.strip() != 'nan')
    counts = df[valid][tracking_col].value_counts()
    dup_tracking = set(counts[counts > 1].index)

    for idx, row in df[valid].iterrows():
        t = row[tracking_col]
        if t in dup_tracking:
            files = df[df[tracking_col] == t]['_source_file'].tolist()
            df.at[idx, '_possible_duplicate'] = 'POSSIBLE DUPLICATE'
            df.at[idx, '_dup_files'] = ', '.join(files)

    # Build duplicates summary dataframe for the Excel tab
    dup_rows = df[df['_possible_duplicate'] == 'POSSIBLE DUPLICATE'].copy()
    if not dup_rows.empty:
        dup_summary = dup_rows[[tracking_col, ref_col, cost_col, date_col, '_source_file', '_dup_files']].copy()
        dup_summary.columns = ['tracking', 'raw_ref', 'cost', 'date', 'source_file', 'appears_in']
        dup_summary['vendor'] = label
    else:
        dup_summary = pd.DataFrame(columns=['tracking','raw_ref','cost','date','source_file','appears_in','vendor'])

    # Aggregate after flagging — group by key, keep duplicate flag if any row in group is flagged
    return df, dup_summary

def agg_nin(df):
    return df.groupby('key').agg(
        nin_invoice_num=('_inv_num', 'first'),
        nin_actual_cost=('_amount', 'sum'),
        nin_shipments=('_order', 'count'),
        nin_raw_ref=('_auth', 'first'),
        nin_date=('_date', 'first')
    ).reset_index()

def agg_wwe(df):
    return df.groupby('key').agg(
        wwe_invoice_num=('_inv_num', 'first'),
        wwe_actual_cost=('_amount', 'sum'),
        wwe_shipments=('_airbill', 'count'),
        wwe_raw_ref=('_ref', 'first'),
        wwe_date=('_date', 'first')
    ).reset_index()

def agg_fedex(df):
    return df.groupby('key').agg(
        fedex_invoice_num=('_inv_num', 'first'),
        fedex_actual_cost=('_amount', 'sum'),
        fedex_shipments=('_tracking', 'count'),
        fedex_raw_ref=('_ref', 'first'),
        fedex_date=('_date', 'first')
    ).reset_index()

def assign_status(row):
    has_wwe    = pd.notna(row.get('wwe_actual_cost'))
    has_nin    = pd.notna(row.get('nin_actual_cost'))
    has_fedex  = pd.notna(row.get('fedex_actual_cost'))
    has_gelato = pd.notna(row.get('gelato_actual_cost'))
    has_vis    = pd.notna(row.get('vision_cost')) and row.get('vision_cost', 0) != 0
    if not has_vis and (has_wwe or has_nin or has_fedex or has_gelato):
        return "NOT IN VISION"
    if has_wwe    and pd.notna(row.get('wwe_diff'))    and abs(row['wwe_diff'])    > 0.01:
        return "MISMATCH"
    if has_nin    and pd.notna(row.get('nin_diff'))    and abs(row['nin_diff'])    > 0.01:
        return "MISMATCH"
    if has_fedex  and pd.notna(row.get('fedex_diff'))  and abs(row['fedex_diff'])  > 0.01:
        return "MISMATCH"
    if has_gelato and pd.notna(row.get('gelato_diff')) and abs(row['gelato_diff']) > 0.01:
        return "MISMATCH"
    if (has_wwe or has_nin or has_fedex or has_gelato) and has_vis:
        return "MATCH"
    return "VISION ONLY"

# --- Excel styling constants ---
ORANGE    = "E8601C"
DARK      = "2B2B2B"
WHITE     = "FFFFFF"
RED_FILL  = "FFCCCC"
GREEN_FILL= "CCFFCC"
YELLOW_FILL="FFF2CC"
GRAY_FILL = "F5F5F5"
MID_GRAY  = "D9D9D9"
LIGHT_BLUE= "E8F4FD"

def tb():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(cell, bg=ORANGE):
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = tb()

def body(cell, center=False, fill=None, bold=False):
    cell.font = Font(name="Arial", bold=bold, size=10)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)
    cell.border = tb()
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)

def label_cell(ws, row, col, text, bg=GRAY_FILL, bold=True, color=DARK):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Arial", bold=bold, size=10, color=color)
    c.fill = PatternFill("solid", fgColor=bg)
    c.border = tb()
    c.alignment = Alignment(horizontal="left", vertical="center")
    return c

def value_cell(ws, row, col, val, bold=False, color=DARK, fill=None):
    c = ws.cell(row=row, column=col, value=round(val, 2) if isinstance(val, float) else val)
    c.font = Font(name="Arial", bold=bold, size=10, color=color)
    c.number_format = '$#,##0.00'
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = tb()
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    return c

def write_detail_row(ws, row_n, raw_ref, rep, dt, desc, vendor, vcost, fill):
    """Write a single indented source line, grouped so it collapses under the summary."""
    ws.row_dimensions[row_n].height = 15
    ws.row_dimensions[row_n].outline_level = 1
    ws.row_dimensions[row_n].hidden = True  # collapsed by default
    vals   = ["  ↳", str(raw_ref) if raw_ref else "", str(rep) if rep else "",
              str(dt) if dt else "", str(desc) if desc else "",
              str(vendor) if vendor else "", vcost, "—", "—", "—", "—"]
    centers = [True, False, False, True, False, False, True, True, True, True, True]
    colors  = ["993C1D","555555","555555","555555","555555","555555","000000","AAAAAA","AAAAAA","AAAAAA","AAAAAA"]
    for ci, (val, cen, col) in enumerate(zip(vals, centers, colors), 1):
        c = ws.cell(row=row_n, column=ci, value=val)
        c.font = Font(name="Arial", size=9, color=col, italic=(ci == 1))
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center" if cen else "left", vertical="center")
        c.border = tb()
        if ci == 7 and isinstance(val, (int, float)):
            c.number_format = '$#,##0.00'

def build_excel(merged, has_nin, has_wwe, has_fedex, has_gelato, vendors_label, period_label, dup_df, vision_raw):
    mismatches    = merged[merged['status'] == 'MISMATCH'].copy()
    matches       = merged[merged['status'] == 'MATCH'].copy()
    vision_only   = merged[merged['status'] == 'VISION ONLY'].copy()
    not_in_vision = merged[merged['status'] == 'NOT IN VISION'].copy()

    # Pre-compute financial data
    all_vendor_rows = merged[merged['status'].isin(['MATCH', 'MISMATCH'])].copy()

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # ── Title banner ──
    ws.merge_cells('A1:G1')
    ws['A1'] = "GSB Digital — Shipping Reconciliation Report"
    ws['A1'].font = Font(name="Arial", bold=True, size=16, color=WHITE)
    ws['A1'].fill = PatternFill("solid", fgColor=ORANGE)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # ── Run info row ──
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Period: {period_label}   |   Vendors included: {vendors_label}   |   Run date: {date.today().strftime('%B %d, %Y')}"
    ws['A2'].font = Font(name="Arial", italic=True, size=9, color="555555")
    ws['A2'].fill = PatternFill("solid", fgColor="F2F2F2")
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A2'].border = tb()
    ws.row_dimensions[2].height = 18

    # ── Stat cards ──
    stats = [
        ("MISMATCHES",    len(mismatches),    "FFCCCC"),
        ("MATCHED",       len(matches),       "CCFFCC"),
        ("NOT IN VISION", len(not_in_vision), "FFF2CC"),
        ("VISION ONLY",   len(vision_only),   LIGHT_BLUE),
    ]
    ws.row_dimensions[4].height = 50
    ws.row_dimensions[5].height = 25
    for i, (label, val, bg) in enumerate(stats, 1):
        c1 = ws.cell(row=4, column=i, value=str(val))
        c1.font = Font(name="Arial", bold=True, size=22)
        c1.fill = PatternFill("solid", fgColor=bg)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.border = tb()
        c2 = ws.cell(row=5, column=i, value=label)
        c2.font = Font(name="Arial", bold=True, size=9, color="555555")
        c2.fill = PatternFill("solid", fgColor=bg)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = tb()

    # ── Section helper ──
    def section_header(ws, row, text, bg=DARK):
        ws.merge_cells(f'A{row}:G{row}')
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name="Arial", bold=True, size=11, color=WHITE)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22

    def col_header_row(ws, row, labels, bg="455A64"):
        for i, lbl in enumerate(labels, 1):
            c = ws.cell(row=row, column=i, value=lbl)
            c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = tb()
        ws.row_dimensions[row].height = 18

    # ── Section 1: ALL invoices total spend ──
    section_header(ws, 7, "TOTAL SPEND — All Matched & Reconciled Invoices")

    vendor_cols = []
    if has_nin:    vendor_cols.append(('NIN (Courier)',  'nin_actual_cost',    'nin_diff'))
    if has_wwe:    vendor_cols.append(('WWE / UPS',      'wwe_actual_cost',    'wwe_diff'))
    if has_fedex:  vendor_cols.append(('FedEx',          'fedex_actual_cost',  'fedex_diff'))
    if has_gelato: vendor_cols.append(('Gelato',         'gelato_actual_cost', 'gelato_diff'))

    col_labels = [""] + [v[0] for v in vendor_cols]
    if has_nin and has_wwe:
        col_labels.append("Combined Total")
    col_header_row(ws, 8, col_labels)

    spend_rows = [
        ("Total Vendor Cost (all invoices)",   'actual', None),
        ("Total Recorded in Vision (all)",      'vision', None),
        ("Net Difference (all invoices)",        'diff',  None),
    ]

    for r_offset, (row_label, metric, _) in enumerate(spend_rows, 9):
        label_cell(ws, r_offset, 1, row_label)
        total = 0
        for col_i, (vname, cost_col, diff_col) in enumerate(vendor_cols, 2):
            subset = all_vendor_rows[all_vendor_rows[cost_col].notna()]
            if metric == 'actual':
                val = subset[cost_col].sum()
            elif metric == 'vision':
                val = subset['vision_cost'].sum()
            else:
                val = subset[diff_col].sum()
            total += val
            is_diff = (metric == 'diff')
            color = ("CC0000" if val < 0 else ("006600" if val > 0 else DARK)) if is_diff else DARK
            value_cell(ws, r_offset, col_i, val, bold=is_diff, color=color)
        if len(vendor_cols) > 1:
            is_diff = (metric == 'diff')
            color = ("CC0000" if total < 0 else ("006600" if total > 0 else DARK)) if is_diff else DARK
            value_cell(ws, r_offset, len(vendor_cols) + 2, total, bold=is_diff, color=color)
        ws.row_dimensions[r_offset].height = 18

    # ── Section 2: Mismatches only ──
    section_row = 9 + len(spend_rows) + 1
    section_header(ws, section_row, "MISMATCH IMPACT — Invoices Where Vision and Vendor Costs Differ", bg="B85C00")

    col_labels = [""] + [v[0] for v in vendor_cols]
    if len(vendor_cols) > 1:
        col_labels.append("Combined Total")
    col_header_row(ws, section_row + 1, col_labels, bg="B85C00")

    mismatch_metric_rows = [
        ("Total Vendor Cost (mismatches only)",  'actual'),
        ("Total Recorded in Vision (mismatches)", 'vision'),
        ("Total Discrepancy (mismatches only)",   'diff'),
    ]

    for r_offset, (row_label, metric) in enumerate(mismatch_metric_rows, section_row + 2):
        label_cell(ws, r_offset, 1, row_label)
        total = 0
        for col_i, (vname, cost_col, diff_col) in enumerate(vendor_cols, 2):
            subset = mismatches[mismatches[cost_col].notna()]
            if metric == 'actual':
                val = subset[cost_col].sum()
            elif metric == 'vision':
                val = subset['vision_cost'].sum()
            else:
                val = subset[diff_col].sum()
            total += val
            is_diff = (metric == 'diff')
            color = ("CC0000" if val < 0 else ("006600" if val > 0 else DARK)) if is_diff else DARK
            fill = "FFF0E8" if is_diff else None
            value_cell(ws, r_offset, col_i, val, bold=is_diff, color=color, fill=fill)
        if len(vendor_cols) > 1:
            is_diff = (metric == 'diff')
            color = ("CC0000" if total < 0 else ("006600" if total > 0 else DARK)) if is_diff else DARK
            fill = "FFF0E8" if is_diff else None
            value_cell(ws, r_offset, len(vendor_cols) + 2, total, bold=is_diff, color=color, fill=fill)
        ws.row_dimensions[r_offset].height = 18

    for i, w in enumerate([36, 18, 18, 18, 5, 5, 5], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ══ Mismatches tab ══
    ws2 = wb.create_sheet("Mismatches")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.outlinePr.summaryBelow = False
    ws2.merge_cells('A1:K1')
    ws2['A1'] = "Shipping Cost Mismatches — Action Required"
    ws2['A1'].font = Font(name="Arial", bold=True, size=13, color=WHITE)
    ws2['A1'].fill = PatternFill("solid", fgColor=ORANGE)
    ws2['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 35
    for i, h in enumerate(["GSB Invoice #","Raw Invoice #","Sales Rep","Date","Description","Vendor",
                            "Vision Cost","Vendor Cost","Difference ($)","Vendor Invoice #","# Shipments"], 1):
        hdr(ws2.cell(row=2, column=i, value=h))
    for i, w in enumerate([13,16,11,10,28,14,13,13,13,16,10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[2].height = 20

    row_n = 3
    for _, r in mismatches.sort_values('vision_invoice').iterrows():
        vendor_pairs = []
        if has_wwe:    vendor_pairs.append(("WWE (UPS)",    r.get('wwe_actual_cost'),    r.get('wwe_diff'),    r.get('wwe_invoice_num',''),    r.get('wwe_shipments')))
        if has_nin:    vendor_pairs.append(("NIN (Courier)",r.get('nin_actual_cost'),    r.get('nin_diff'),    r.get('nin_invoice_num',''),    r.get('nin_shipments')))
        if has_fedex:  vendor_pairs.append(("FedEx",        r.get('fedex_actual_cost'),  r.get('fedex_diff'),  r.get('fedex_invoice_num',''),  r.get('fedex_shipments')))
        if has_gelato: vendor_pairs.append(("Gelato",       r.get('gelato_actual_cost'), r.get('gelato_diff'), r.get('gelato_carrier',''),     r.get('gelato_shipments')))
        for vendor, cost, diff, inv, ships in vendor_pairs:
            if pd.notna(cost):
                fill = RED_FILL if diff < -0.01 else (YELLOW_FILL if diff > 0.01 else GREEN_FILL)
                detail_fill = "FFF5F5" if diff < -0.01 else ("FFFBF0" if diff > 0.01 else "F2FFF2")
                # Get source lines for this invoice key
                src_lines = vision_raw[vision_raw['key'] == r['key']]
                n_lines = len(src_lines)
                # Summary row
                row_data = [r['vision_invoice'], r.get('vision_invoice_raw',''), r['vision_sales_rep'],
                            r['vision_date'], r['vision_description'], vendor,
                            r['vision_cost'], cost, round(diff, 2), inv,
                            int(ships) if pd.notna(ships) else '']
                ws2.row_dimensions[row_n].height = 18
                for ci, val in enumerate(row_data, 1):
                    c = ws2.cell(row=row_n, column=ci, value=val)
                    diff_color = "A32D2D" if diff < -0.01 else ("27500A" if diff > 0.01 else "000000")
                    c.font = Font(name="Arial", bold=(ci in [7,8,9]), size=10,
                                  color=diff_color if ci == 9 else "000000")
                    c.fill = PatternFill("solid", fgColor=fill)
                    c.alignment = Alignment(horizontal="center" if ci in [1,2,4,6,7,8,9,10,11] else "left",
                                            vertical="center")
                    c.border = tb()
                    if ci in [7, 8, 9]:
                        c.number_format = '$#,##0.00'
                row_n += 1
                # Detail rows — one per Vision source line
                for _, src in src_lines.iterrows():
                    write_detail_row(ws2, row_n,
                                     src['Invoice'], src['Sales Rep'], src['Pickup Date'],
                                     src['Description'], "", src['Cost'], detail_fill)
                    row_n += 1

    ws2.row_dimensions[row_n].height = 22
    ws2.cell(row=row_n, column=6, value="TOTALS").font = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=row_n, column=6).alignment = Alignment(horizontal="right")
    for ci in [7, 8, 9]:
        cl = get_column_letter(ci)
        c = ws2.cell(row=row_n, column=ci, value=f'=SUM({cl}3:{cl}{row_n-1})')
        c.font = Font(name="Arial", bold=True, size=10)
        c.number_format = '$#,##0.00'
        c.fill = PatternFill("solid", fgColor=MID_GRAY)
        c.border = tb()
        c.alignment = Alignment(horizontal="center")

    # ══ Matched tab ══
    ws3 = wb.create_sheet("Matched")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.outlinePr.summaryBelow = False
    ws3.merge_cells('A1:I1')
    ws3['A1'] = "Verified Matches — No Action Required"
    ws3['A1'].font = Font(name="Arial", bold=True, size=13, color=WHITE)
    ws3['A1'].fill = PatternFill("solid", fgColor="2E7D32")
    ws3['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 35
    for i, h in enumerate(["GSB Invoice #","Raw Invoice #","Sales Rep","Date","Vendor","Vendor Invoice #",
                            "Vision Cost","Vendor Cost","Difference"], 1):
        hdr(ws3.cell(row=2, column=i, value=h), bg="2E7D32")
    for i, w in enumerate([13,16,11,10,14,16,13,13,12], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.row_dimensions[2].height = 20
    row_n3 = 3
    for _, r in matches.sort_values('vision_invoice').iterrows():
        vendor_pairs = []
        if has_wwe:    vendor_pairs.append(("WWE (UPS)",    r.get('wwe_actual_cost'),    r.get('wwe_diff')))
        if has_nin:    vendor_pairs.append(("NIN (Courier)",r.get('nin_actual_cost'),    r.get('nin_diff')))
        if has_fedex:  vendor_pairs.append(("FedEx",        r.get('fedex_actual_cost'),  r.get('fedex_diff')))
        if has_gelato: vendor_pairs.append(("Gelato",       r.get('gelato_actual_cost'), r.get('gelato_diff')))
        for vendor, cost, diff in vendor_pairs:
            if pd.notna(cost):
                src_lines = vision_raw[vision_raw['key'] == r['key']]
                # Summary row
                ws3.row_dimensions[row_n3].height = 18
                for ci, val in enumerate([r['vision_invoice'], r.get('vision_invoice_raw',''),
                                          r['vision_sales_rep'], r['vision_date'],
                                          vendor, '', r['vision_cost'], cost, round(diff, 2)], 1):
                    c = ws3.cell(row=row_n3, column=ci, value=val)
                    body(c, center=(ci in [1,2,4,5,6,7,8,9]), fill=GREEN_FILL)
                    if ci in [7, 8, 9]:
                        c.number_format = '$#,##0.00'
                row_n3 += 1
                # Detail rows
                for _, src in src_lines.iterrows():
                    write_detail_row(ws3, row_n3,
                                     src['Invoice'], src['Sales Rep'], src['Pickup Date'],
                                     src['Description'], "", src['Cost'], "F2FFF2")
                    row_n3 += 1

    # ══ Not in Vision tab ══
    ws4 = wb.create_sheet("Not in Vision")
    ws4.sheet_view.showGridLines = False

    # Split into two groups:
    # - Rows with a real GSB invoice key (no Vision match found)
    # - Rows flagged as UNMATCHED (missing or unreadable reference — needs manual review)
    unmatched_pattern = re.compile(r'^(NIN|WWE|FEDEX)_UNMATCHED_\d+$')

    def write_niv_section(ws, start_row, section_title, section_color, rows_df, has_nin, has_wwe, has_fedex):
        # Section header
        ws.merge_cells(f'A{start_row}:H{start_row}')
        c = ws.cell(row=start_row, column=1, value=section_title)
        c.font = Font(name="Arial", bold=True, size=12, color=WHITE)
        c.fill = PatternFill("solid", fgColor=section_color)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[start_row].height = 28

        # Column headers
        col_headers = ["GSB Invoice # / Key", "Original Vendor Reference",
                       "Vendor", "Vendor Invoice #", "# Shipments",
                       "Vendor Cost", "Notes", "Resolved?"]
        for i, h in enumerate(col_headers, 1):
            hdr(ws.cell(row=start_row+1, column=i, value=h), bg=section_color)
        ws.row_dimensions[start_row+1].height = 18

        row_n = start_row + 2
        for _, r in rows_df.sort_values('key').iterrows():
            vendor_pairs = []
            if has_wwe:    vendor_pairs.append(("WWE (UPS)",    r.get('wwe_actual_cost'),    r.get('wwe_shipments'),    r.get('wwe_invoice_num',''),    r.get('wwe_raw_ref','')))
            if has_nin:    vendor_pairs.append(("NIN (Courier)",r.get('nin_actual_cost'),    r.get('nin_shipments'),    r.get('nin_invoice_num',''),    r.get('nin_raw_ref','')))
            if has_fedex:  vendor_pairs.append(("FedEx",        r.get('fedex_actual_cost'),  r.get('fedex_shipments'),  r.get('fedex_invoice_num',''),  r.get('fedex_raw_ref','')))
            if has_gelato: vendor_pairs.append(("Gelato",       r.get('gelato_actual_cost'), r.get('gelato_shipments'), r.get('gelato_carrier',''),     r.get('gelato_raw_ref','')))
            for vendor, cost, ships, inv, raw_ref in vendor_pairs:
                if pd.notna(cost):
                    ws.row_dimensions[row_n].height = 18
                    display_key = raw_ref if (pd.isna(r['key']) or unmatched_pattern.match(str(r['key']))) else r['key']
                    display_raw = raw_ref if raw_ref else '— no reference —'
                    note = "⚠ No usable reference — manual review required" if unmatched_pattern.match(str(r['key'])) else "Invoice not matched in Vision"
                    row_data = [display_key, display_raw, vendor, inv,
                                int(ships) if pd.notna(ships) else '',
                                cost, note, ""]
                    for ci, val in enumerate(row_data, 1):
                        c = ws.cell(row=row_n, column=ci, value=val)
                        fill = "FFE0E0" if unmatched_pattern.match(str(r['key'])) else YELLOW_FILL
                        body(c, center=(ci in [3,4,5,6,8]),
                             fill=fill if ci in [1,2,6,7] else None)
                        if ci == 6:
                            c.number_format = '$#,##0.00'
                    row_n += 1
        return row_n  # return next available row

    for i, w in enumerate([22, 28, 14, 18, 12, 14, 40, 12], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # Title banner
    ws4.merge_cells('A1:H1')
    ws4['A1'] = "Vendor Charges Not Found in Vision — Review Required"
    ws4['A1'].font = Font(name="Arial", bold=True, size=13, color=WHITE)
    ws4['A1'].fill = PatternFill("solid", fgColor="B8860B")
    ws4['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 35

    # Section 1 — rows with a real reference that just didn't match Vision
    matched_ref_rows = not_in_vision[~not_in_vision['key'].apply(
        lambda k: bool(unmatched_pattern.match(str(k))))]
    # Section 2 — rows with no usable reference at all
    unmatched_ref_rows = not_in_vision[not_in_vision['key'].apply(
        lambda k: bool(unmatched_pattern.match(str(k))))]

    next_row = write_niv_section(ws4, 3,
        "SECTION 1 — Charges with a Reference Number (not matched to Vision)",
        "B8860B", matched_ref_rows, has_nin, has_wwe, has_fedex)

    if not unmatched_ref_rows.empty:
        next_row += 1  # blank spacer row
        write_niv_section(ws4, next_row,
            "SECTION 2 — Charges with NO Reference — Manual Review Required",
            "CC0000", unmatched_ref_rows, has_nin, has_wwe, has_fedex)

    # ══ Vision Only tab ══
    ws5 = wb.create_sheet("Vision Only")
    ws5.sheet_view.showGridLines = False
    ws5.merge_cells('A1:G1')
    ws5['A1'] = "Vision Entries with No Vendor Invoice — For Reference"
    ws5['A1'].font = Font(name="Arial", bold=True, size=13, color=WHITE)
    ws5['A1'].fill = PatternFill("solid", fgColor="455A64")
    ws5['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 35
    for i, h in enumerate(["GSB Invoice #","Raw Invoice #","Sales Rep","Date","Description",
                            "Amount Billed","Cost Recorded","Notes"], 1):
        hdr(ws5.cell(row=2, column=i, value=h), bg="455A64")
    for i, w in enumerate([14,16,12,10,36,14,14,28], 1):
        ws5.column_dimensions[get_column_letter(i)].width = w
    row_n5 = 3
    for _, r in vision_only.sort_values('vision_invoice').iterrows():
        ws5.row_dimensions[row_n5].height = 18
        for ci, val in enumerate([r['vision_invoice'], r.get('vision_invoice_raw',''),
                                   r['vision_sales_rep'], r['vision_date'],
                                   r['vision_description'], r['vision_billed'], r['vision_cost'], ""], 1):
            c = ws5.cell(row=row_n5, column=ci, value=val)
            body(c, center=(ci in [1,2,3,4,6,7]))
            if ci in [6, 7]:
                c.number_format = '$#,##0.00'
        row_n5 += 1

    # ══ Possible Duplicates tab ══
    ws6 = wb.create_sheet("Possible Duplicates")
    ws6.sheet_view.showGridLines = False

    # Title banner
    ws6.merge_cells('A1:H1')
    ws6['A1'] = "Possible Duplicates — Review Required"
    ws6['A1'].font = Font(name="Arial", bold=True, size=13, color=WHITE)
    ws6['A1'].fill = PatternFill("solid", fgColor=ORANGE)
    ws6['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws6.row_dimensions[1].height = 35

    # Info row
    ws6.merge_cells('A2:H2')
    ws6['A2'] = "These rows appear in more than one uploaded file for the same vendor. Nothing has been deleted. All rows still appear in every other tab. Your team must review and mark resolved."
    ws6['A2'].font = Font(name="Arial", italic=True, size=9, color="555555")
    ws6['A2'].fill = PatternFill("solid", fgColor="F2F2F2")
    ws6['A2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws6['A2'].border = tb()
    ws6.row_dimensions[2].height = 28

    for i, w in enumerate([18, 22, 14, 12, 16, 14, 40, 14], 1):
        ws6.column_dimensions[get_column_letter(i)].width = w

    if dup_df.empty:
        ws6.merge_cells('A4:H4')
        c = ws6.cell(row=4, column=1, value="No duplicate tracking numbers detected across uploaded files.")
        c.font = Font(name="Arial", italic=True, size=10, color="555555")
        c.fill = PatternFill("solid", fgColor=GREEN_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws6.row_dimensions[4].height = 28
    else:
        # Grand total row
        total_flagged = len(dup_df)
        total_at_risk = dup_df['cost'].sum()
        vendors_affected = dup_df['vendor'].nunique()
        ws6.row_dimensions[4].height = 35
        for ci, (label_t, val_t) in enumerate([
            ("TOTAL FLAGGED ROWS", total_flagged),
            ("TOTAL $ FLAGGED FOR REVIEW", f"${total_at_risk:,.2f}"),
            ("VENDORS AFFECTED", vendors_affected)
        ], 1):
            c_top = ws6.cell(row=4, column=ci, value=str(val_t))
            c_top.font = Font(name="Arial", bold=True, size=16)
            c_top.fill = PatternFill("solid", fgColor="FFF2CC")
            c_top.alignment = Alignment(horizontal="center", vertical="center")
            c_top.border = tb()
            c_bot = ws6.cell(row=5, column=ci, value=label_t)
            c_bot.font = Font(name="Arial", bold=True, size=9, color="555555")
            c_bot.fill = PatternFill("solid", fgColor="FFF2CC")
            c_bot.alignment = Alignment(horizontal="center", vertical="center")
            c_bot.border = tb()
        ws6.row_dimensions[5].height = 18

        current_row = 7
        for vendor_label in dup_df['vendor'].unique():
            vendor_dups = dup_df[dup_df['vendor'] == vendor_label].copy()
            vendor_total = vendor_dups['cost'].sum()

            # Section header
            ws6.merge_cells(f'A{current_row}:H{current_row}')
            c = ws6.cell(row=current_row, column=1,
                         value=f"{vendor_label}  —  {len(vendor_dups)} flagged rows  —  ${vendor_total:,.2f} flagged for review")
            c.font = Font(name="Arial", bold=True, size=11, color=WHITE)
            c.fill = PatternFill("solid", fgColor=DARK)
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws6.row_dimensions[current_row].height = 22
            current_row += 1

            # Column headers
            for ci, h in enumerate(["GSB Invoice # / Key", "Tracking #", "Date",
                                     "Amount", "Source File", "Also Appears In",
                                     "Flag", "Resolved?"], 1):
                hdr(ws6.cell(row=current_row, column=ci, value=h), bg="555555")
            ws6.row_dimensions[current_row].height = 18
            current_row += 1

            for _, dr in vendor_dups.sort_values('tracking').iterrows():
                ws6.row_dimensions[current_row].height = 18
                row_vals = [
                    dr.get('raw_ref', ''),
                    dr.get('tracking', ''),
                    dr.get('date', ''),
                    dr.get('cost', 0),
                    dr.get('source_file', ''),
                    dr.get('appears_in', ''),
                    '⚠ Possible duplicate',
                    ''
                ]
                for ci, val in enumerate(row_vals, 1):
                    c = ws6.cell(row=current_row, column=ci, value=val)
                    body(c, center=(ci in [2, 3, 4, 7, 8]), fill=YELLOW_FILL if ci in [4, 7] else None)
                    if ci == 4:
                        c.number_format = '$#,##0.00'
                current_row += 1

            # Vendor totals row
            ws6.row_dimensions[current_row].height = 20
            c_lbl = ws6.cell(row=current_row, column=3, value="TOTALS")
            c_lbl.font = Font(name="Arial", bold=True, size=10)
            c_lbl.alignment = Alignment(horizontal="right")
            c_lbl.fill = PatternFill("solid", fgColor=MID_GRAY)
            c_lbl.border = tb()
            c_tot = ws6.cell(row=current_row, column=4, value=round(vendor_total, 2))
            c_tot.font = Font(name="Arial", bold=True, size=10)
            c_tot.number_format = '$#,##0.00'
            c_tot.alignment = Alignment(horizontal="center")
            c_tot.fill = PatternFill("solid", fgColor=MID_GRAY)
            c_tot.border = tb()
            lbl2 = ws6.cell(row=current_row, column=5,
                            value=f"{len(vendor_dups)} rows flagged across uploaded files")
            lbl2.font = Font(name="Arial", size=9, color="555555")
            lbl2.fill = PatternFill("solid", fgColor=MID_GRAY)
            lbl2.border = tb()
            current_row += 2  # blank spacer between vendors

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Ready check ──
ready = vision_file and len(vendor_files) > 0

if st.button("Run Reconciliation", disabled=not ready):
    with st.spinner("Matching invoices and calculating discrepancies..."):
        try:
            # Route each uploaded vendor file to the correct bucket by reading its columns
            nin_files    = []
            wwe_files    = []
            fedex_files  = []
            gelato_files = []
            unknown_files = []

            for f in vendor_files:
                vendor_type = detect_vendor(f)
                if vendor_type == 'nin':        nin_files.append(f)
                elif vendor_type == 'wwe':      wwe_files.append(f)
                elif vendor_type == 'fedex':    fedex_files.append(f)
                elif vendor_type == 'gelato':   gelato_files.append(f)
                else:                           unknown_files.append(f)

            if unknown_files:
                for uf in unknown_files:
                    st.warning(
                        f"⚠ **Could not identify:** `{uf.name}` — this file was skipped. "
                        f"The tool supports NIN (.xls/.xlsx), WWE (.xls/.xlsx), "
                        f"FedEx (.csv/.xls/.xlsx), and Gelato (.xlsx/.csv) exports. "
                        f"If this is a supported file, the column names may have changed — "
                        f"contact whoever maintains this tool."
                    )

            has_nin    = len(nin_files) > 0
            has_wwe    = len(wwe_files) > 0
            has_fedex  = len(fedex_files) > 0
            has_gelato = len(gelato_files) > 0

            if not (has_nin or has_wwe or has_fedex or has_gelato):
                raise ReconError(
                    code="RUN-001",
                    title="No recognizable vendor files uploaded",
                    detail="None of the uploaded files could be identified as a NIN, WWE, FedEx, or Gelato invoice.",
                    fix="Check that you're uploading the correct vendor invoice files. "
                        "Supported formats: NIN (.xls/.xlsx), WWE (.xls/.xlsx), "
                        "FedEx (.csv/.xls/.xlsx), Gelato (.xlsx/.csv). "
                        "If a file is being skipped, check the warning message above it for details."
                )

            # Show what was detected
            detected = []
            if has_nin:    detected.append(f"NIN ({len(nin_files)} file{'s' if len(nin_files)>1 else ''})")
            if has_wwe:    detected.append(f"WWE ({len(wwe_files)} file{'s' if len(wwe_files)>1 else ''})")
            if has_fedex:  detected.append(f"FedEx ({len(fedex_files)} file{'s' if len(fedex_files)>1 else ''})")
            if has_gelato: detected.append(f"Gelato ({len(gelato_files)} file{'s' if len(gelato_files)>1 else ''})")
            st.info(f"Detected: {', '.join(detected)}")

            vision_agg, vision_raw = load_vision(vision_file)
            merged = vision_agg.copy()
            all_dup_frames = []

            if has_wwe:
                wwe_combined = combine_vendor_files(wwe_files, load_wwe)
                if wwe_combined is None:
                    raise ReconError("WWE-002", "No WWE data loaded",
                                     "All WWE files failed to load.",
                                     "Check the WWE files and try again.")
                wwe_combined, wwe_dups = flag_duplicates(
                    wwe_combined, '_tracking', '_amount', '_date',
                    '_ref', 'WWE / UPS')
                if not wwe_dups.empty: all_dup_frames.append(wwe_dups)
                wwe_agg = agg_wwe(wwe_combined)
                merged = merged.merge(wwe_agg, on='key', how='outer')

            if has_nin:
                nin_combined = combine_vendor_files(nin_files, load_nin)
                if nin_combined is None:
                    raise ReconError("NIN-002", "No NIN data loaded",
                                     "All NIN files failed to load.",
                                     "Check the NIN files and try again.")
                nin_combined, nin_dups = flag_duplicates(
                    nin_combined, '_tracking', '_amount', '_date',
                    '_auth', 'NIN (Courier)')
                if not nin_dups.empty: all_dup_frames.append(nin_dups)
                nin_agg = agg_nin(nin_combined)
                merged = merged.merge(nin_agg, on='key', how='outer')

            if has_fedex:
                fedex_combined = combine_vendor_files(fedex_files, load_fedex)
                if fedex_combined is None:
                    raise ReconError("FDX-002", "No FedEx data loaded",
                                     "All FedEx files failed to load.",
                                     "Check the FedEx files and try again.")
                fedex_combined, fedex_dups = flag_duplicates(
                    fedex_combined, '_tracking', '_amount', '_date',
                    '_ref', 'FedEx')
                if not fedex_dups.empty: all_dup_frames.append(fedex_dups)
                fedex_agg = agg_fedex(fedex_combined)
                merged = merged.merge(fedex_agg, on='key', how='outer')

            if has_gelato:
                gelato_combined = combine_vendor_files(gelato_files, load_gelato)
                if gelato_combined is None:
                    raise ReconError("GEL-004", "No Gelato data loaded",
                                     "All Gelato files failed to load.",
                                     "Check the Gelato files and try again.")
                gelato_combined, gelato_dups = flag_duplicates(
                    gelato_combined, '_tracking', '_cost', '_date',
                    '_ref', 'Gelato')
                if not gelato_dups.empty: all_dup_frames.append(gelato_dups)
                gelato_agg = agg_gelato(gelato_combined)
                merged = merged.merge(gelato_agg, on='key', how='outer')

            # Never drop rows — only remove phantom rows with no cost data at all
            cost_cols = [c for c in ['vision_cost','wwe_actual_cost','nin_actual_cost',
                                     'fedex_actual_cost','gelato_actual_cost']
                         if c in merged.columns]
            has_any_value = merged[cost_cols].notna().any(axis=1)
            merged = merged[has_any_value].copy()

            # Sanity check — make sure the merge produced something
            if merged.empty:
                raise ReconError(
                    code="RUN-002",
                    title="No data after merging files",
                    detail="The Vision export and vendor invoices were loaded but produced no rows after merging. "
                           "This usually means the date ranges don't overlap at all.",
                    fix="Check that the Vision export and vendor invoices cover the same billing period. "
                        "The date ranges for each file are shown after you run the tool."
                )

            if has_wwe:    merged['wwe_diff']    = (merged['wwe_actual_cost']    - merged['vision_cost']).round(2)
            if has_nin:    merged['nin_diff']    = (merged['nin_actual_cost']    - merged['vision_cost']).round(2)
            if has_fedex:  merged['fedex_diff']  = (merged['fedex_actual_cost']  - merged['vision_cost']).round(2)
            if has_gelato: merged['gelato_diff'] = (merged['gelato_actual_cost'] - merged['vision_cost']).round(2)

            merged['status'] = merged.apply(assign_status, axis=1)

            mismatches    = merged[merged['status'] == 'MISMATCH']
            matches       = merged[merged['status'] == 'MATCH']
            not_in_vision = merged[merged['status'] == 'NOT IN VISION']
            vision_only   = merged[merged['status'] == 'VISION ONLY']
            dup_df        = pd.concat(all_dup_frames, ignore_index=True) if all_dup_frames else pd.DataFrame()

            col1, col2, col3, col4, col5 = st.columns(5)
            col1.metric("Mismatches",          len(mismatches))
            col2.metric("Matched",             len(matches))
            col3.metric("Not in Vision",       len(not_in_vision))
            col4.metric("Vision Only",         len(vision_only))
            col5.metric("Possible Duplicates", len(dup_df))

            vendor_parts = []
            if has_nin:    vendor_parts.append("NIN")
            if has_wwe:    vendor_parts.append("WWE")
            if has_fedex:  vendor_parts.append("FedEx")
            if has_gelato: vendor_parts.append("Gelato")
            vendors_label = " + ".join(vendor_parts)

            if period_start and period_end:
                period_label = f"{period_start.strftime('%b %d %Y')} – {period_end.strftime('%b %d %Y')}"
                period_slug  = f"{period_start.strftime('%Y%m%d')}_to_{period_end.strftime('%Y%m%d')}"
            else:
                period_label = "Period not specified"
                period_slug  = date.today().strftime('%Y%m%d')

            filename  = f"GSB_Shipping_Recon_{period_slug}.xlsx"
            excel_buf = build_excel(merged, has_nin, has_wwe, has_fedex, has_gelato,
                                    vendors_label, period_label, dup_df, vision_raw)
            st.success("Reconciliation complete. Download your report below.")
            st.download_button(
                label="Download Reconciliation Report (.xlsx)",
                data=excel_buf,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        except ReconError as e:
            show_error(e)
        except Exception as e:
            show_error(e)

elif not ready:
    if not vision_file:
        st.info("Upload the Vision export to get started. Then drop your vendor invoice files in the box below.")
    else:
        st.info("Vision file uploaded. Now drop at least one vendor invoice file to run the reconciliation.")
